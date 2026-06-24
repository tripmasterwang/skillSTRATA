from __future__ import annotations

import datetime
import json
import os

import openpyxl


def datetime_to_float(dt):
    excel_start_date = datetime.datetime(1899, 12, 30)
    delta = dt - excel_start_date
    return delta.days + delta.seconds / 86400.0


def transform_value(value):
    if isinstance(value, (int, float)):
        value = round(float(value), 2)
    elif isinstance(value, datetime.time):
        value = str(value)[:-3]
    elif isinstance(value, datetime.datetime):
        value = round(datetime_to_float(value), 0)
    elif isinstance(value, str):
        try:
            value = round(float(value), 2)
        except ValueError:
            pass
    return value


def compare_cell_value(v1, v2):
    v1 = transform_value(v1)
    v2 = transform_value(v2)
    if (v1 == "" and v2 is None) or (v1 is None and v2 == ""):
        return True
    if (v1 == "" and v2 == "") or (v1 is None and v2 is None):
        return True
    if type(v1) != type(v2):
        return False
    return v1 == v2


def col_num2name(n):
    name = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        name = chr(65 + remainder) + name
    return name


def col_name2num(name):
    num = 0
    for char in name:
        num = num * 26 + (ord(char) - ord("A") + 1)
    return num


def parse_cell_range(range_str):
    start_cell, end_cell = range_str.split(":")
    start_col, start_row = "", ""
    for char in start_cell:
        if char.isdigit():
            start_row += char
        else:
            start_col += char

    end_col, end_row = "", ""
    for char in end_cell:
        if char.isdigit():
            end_row += char
        else:
            end_col += char

    return start_col, start_row, end_col, end_row


def generate_cell_names(range_str, max_row=None, max_col=None):
    if ":" not in range_str:
        return [range_str]

    start_col, start_row, end_col, end_row = parse_cell_range(range_str)

    if not start_row and not end_row:
        start_row = "1"
        end_row = str(max_row) if max_row else "1000"
    elif not start_row:
        start_row = "1"
    elif not end_row:
        end_row = str(max_row) if max_row else "1000"

    if not start_col and not end_col:
        start_col = "A"
        end_col = col_num2name(max_col) if max_col else "Z"
    elif not start_col:
        start_col = "A"
    elif not end_col:
        end_col = col_num2name(max_col) if max_col else "Z"

    start_col_num = col_name2num(start_col)
    end_col_num = col_name2num(end_col)
    start_row_num = int(start_row)
    end_row_num = int(end_row)

    columns = [col_num2name(i) for i in range(start_col_num, end_col_num + 1)]
    return [f"{col}{row}" for col in columns for row in range(start_row_num, end_row_num + 1)]


def cell_level_compare(wb_gt, wb_proc, sheet_name, cell_range):
    if sheet_name not in wb_proc.sheetnames:
        return False, f"Worksheet '{sheet_name}' not found in output"
    ws_gt = wb_gt[sheet_name]
    ws_proc = wb_proc[sheet_name]
    cell_names = generate_cell_names(cell_range, max_row=ws_gt.max_row, max_col=ws_gt.max_column)

    for cell_name in cell_names:
        cell_gt = ws_gt[cell_name]
        cell_proc = ws_proc[cell_name]
        if not compare_cell_value(cell_gt.value, cell_proc.value):
            msg = f"Value mismatch at {cell_name}: expected '{cell_gt.value}', got '{cell_proc.value}'"
            return False, msg

    return True, ""


def compare_workbooks(gt_file, output_file, answer_position):
    if not os.path.exists(output_file):
        return False, "Output file not found"

    try:
        wb_gt = openpyxl.load_workbook(filename=gt_file, data_only=True)
        wb_proc = openpyxl.load_workbook(filename=output_file, data_only=True)
    except Exception as exc:
        return False, f"Error loading workbook: {exc}"

    for sheet_cell_range in answer_position.split(","):
        sheet_cell_range = sheet_cell_range.strip()
        if "!" in sheet_cell_range:
            sheet_name, cell_range = sheet_cell_range.split("!")
            sheet_name = sheet_name.strip("'")
        else:
            sheet_name = wb_gt.sheetnames[0]
            cell_range = sheet_cell_range

        cell_range = cell_range.strip("'")
        result, message = cell_level_compare(wb_gt, wb_proc, sheet_name, cell_range)
        if not result:
            return False, message

    return True, ""


def load_dataset(data_path):
    candidates = [
        os.path.join(data_path, "dataset.json"),
        os.path.join(data_path, "data.json"),
        os.path.join(data_path, "data.jsonl"),
        os.path.join(data_path, "sample_data.jsonl"),
    ]

    if data_path.endswith(".json") or data_path.endswith(".jsonl"):
        candidates.insert(0, data_path)

    data_file = None
    for path in candidates:
        if os.path.exists(path):
            data_file = path
            break

    if data_file is None and os.path.isdir(data_path):
        for filename in os.listdir(data_path):
            if filename.endswith(".json") or filename.endswith(".jsonl"):
                data_file = os.path.join(data_path, filename)
                break

    if data_file is None:
        raise FileNotFoundError(f"No dataset file found in {data_path}")

    with open(data_file, "r", encoding="utf-8") as handle:
        if data_file.endswith(".json"):
            return json.load(handle)
        return [json.loads(line) for line in handle if line.strip()]


def find_spreadsheet_dir(data_path, instance):
    data_dir = data_path if os.path.isdir(data_path) else os.path.dirname(data_path)
    instance_id = str(instance["id"])
    spreadsheet_path = str(instance.get("spreadsheet_path", instance_id))

    candidates = [
        os.path.join(data_dir, "spreadsheet", spreadsheet_path),
        os.path.join(data_dir, spreadsheet_path),
        os.path.join(data_dir, "spreadsheet", instance_id),
        os.path.join(data_dir, instance_id),
    ]

    for path in candidates:
        if os.path.exists(path):
            return path

    return None


def find_output_dir(output_base, instance):
    instance_id = str(instance["id"])
    spreadsheet_path = str(instance.get("spreadsheet_path", instance_id))

    candidates = [
        os.path.join(output_base, spreadsheet_path),
        os.path.join(output_base, instance_id),
    ]

    for path in candidates:
        if os.path.isdir(path):
            return path

    return candidates[0]
