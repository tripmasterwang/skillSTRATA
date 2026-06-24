"""
SpreadsheetBench Runner - Task execution logic.

This module handles loading benchmark data, running agents on tasks,
and collecting results. It is agnostic to the specific agent implementation.
"""

import json
import os
import shutil
import tempfile
import traceback
from dataclasses import dataclass, field
from datetime import datetime
from typing import Iterator

from .agents.base import BaseSpreadsheetAgent, AgentContext


@dataclass
class BenchmarkInstance:
    """A single benchmark instance."""
    id: str
    instruction: str
    spreadsheet_path: str
    instruction_type: str = ""
    answer_position: str = ""
    metadata: dict = field(default_factory=dict)


@dataclass
class TestCaseResult:
    """Result for a single test case."""
    input_file: str
    output_file: str
    success: bool
    agent_answer: str = ""
    turns: int = 0
    error: str = ""


@dataclass
class InstanceResult:
    """Result for a benchmark instance (may have multiple test cases)."""
    id: str
    instruction: str
    success: bool
    test_cases: list[TestCaseResult] = field(default_factory=list)
    error: str = ""


@dataclass
class BenchmarkResult:
    """Overall benchmark results."""
    agent_name: str
    model: str
    timestamp: str
    total_instances: int
    successful_instances: int
    success_rate: float
    results: list[InstanceResult] = field(default_factory=list)


def get_spreadsheet_content(file_path: str, max_rows: int = 5) -> str:
    """Get spreadsheet content in the format expected by SpreadsheetBench."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        lines = []
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            if i > max_rows:
                lines.append(f"... ({ws.max_row - max_rows} more rows)")
                break
            # Format as tuple-like representation
            row_values = [str(cell) if cell is not None else "" for cell in row]
            lines.append(str(tuple(row_values)))

        wb.close()
        return "\n".join(lines)
    except Exception as e:
        return f"[Could not read spreadsheet: {e}]"


class SpreadsheetBenchRunner:
    """
    Runner for executing agents on SpreadsheetBench.

    Handles:
    - Loading benchmark data from JSONL or JSON
    - Setting up working directories
    - Running agents on each instance
    - Collecting and saving results
    """

    def __init__(
        self,
        agent: BaseSpreadsheetAgent,
        data_path: str,
        output_dir: str = "outputs/spreadsheetbench",
        working_dir: str | None = None,
    ):
        """
        Initialize the runner.

        Args:
            agent: The agent to run
            data_path: Path to benchmark data (directory, JSONL, or JSON file)
            output_dir: Directory to save output spreadsheets
            working_dir: Working directory for agent (default: system temp folder)
        """
        self.agent = agent
        self.data_path = data_path
        self.output_dir = output_dir
        # If a working_dir is provided, initialize it immediately so callers can
        # use run_instance() directly (e.g., parallel workers) without calling run().
        if working_dir is not None:
            working_dir = os.path.abspath(working_dir)
            os.makedirs(working_dir, exist_ok=True)
        self._custom_working_dir = working_dir
        self.working_dir: str | None = working_dir

    def load_data(self) -> list[BenchmarkInstance]:
        """Load benchmark instances from JSONL or JSON file."""
        data_file = self._find_data_file()
        if data_file is None:
            raise FileNotFoundError(f"Could not find data file (JSONL or JSON) in {self.data_path}")

        print(f"Loading data from: {data_file}")
        instances = []

        with open(data_file, "r", encoding="utf-8") as f:
            if data_file.endswith(".json"):
                # JSON array format
                data_list = json.load(f)
            else:
                # JSONL format (one JSON object per line)
                data_list = []
                for line in f:
                    line = line.strip()
                    if line:
                        data_list.append(json.loads(line))

        for data in data_list:
            instance_id = str(data["id"])
            instances.append(BenchmarkInstance(
                id=instance_id,
                instruction=data["instruction"],
                spreadsheet_path=str(data.get("spreadsheet_path", instance_id)),
                instruction_type=data.get("instruction_type", ""),
                answer_position=data.get("answer_position", ""),
                metadata={k: v for k, v in data.items()
                          if k not in ("id", "instruction", "spreadsheet_path",
                                       "instruction_type", "answer_position")},
            ))

        return instances

    def _find_data_file(self) -> str | None:
        """Find the data file (JSONL or JSON)."""
        candidates = [
            # JSONL files
            os.path.join(self.data_path, "data.jsonl"),
            os.path.join(self.data_path, "sample_data.jsonl"),
            self.data_path if self.data_path.endswith(".jsonl") else None,
            # JSON files
            os.path.join(self.data_path, "dataset.json"),
            os.path.join(self.data_path, "data.json"),
            self.data_path if self.data_path.endswith(".json") else None,
        ]

        if os.path.isdir(self.data_path):
            for f in os.listdir(self.data_path):
                if f.endswith(".jsonl") or f.endswith(".json"):
                    candidates.append(os.path.join(self.data_path, f))

        for path in candidates:
            if path and os.path.exists(path):
                return path

        return None

    def _find_spreadsheet_dir(self, instance: BenchmarkInstance) -> str | None:
        """Find the spreadsheet directory for an instance."""
        data_dir = self.data_path if os.path.isdir(self.data_path) else os.path.dirname(self.data_path)

        candidates = [
            os.path.join(data_dir, "spreadsheet", instance.spreadsheet_path),
            os.path.join(data_dir, instance.spreadsheet_path),
            os.path.join(data_dir, "spreadsheet", instance.id),
            os.path.join(data_dir, instance.id),
        ]

        for path in candidates:
            if os.path.exists(path):
                return path

        return None

    def _find_input_files(self, spreadsheet_dir: str) -> list[str]:
        """Find input files in the spreadsheet directory."""
        files = os.listdir(spreadsheet_dir)

        # Try standard naming convention first: *_input.xlsx
        input_files = sorted([
            f for f in files
            if f.endswith("_input.xlsx")
        ])

        if not input_files:
            # Try verified dataset naming: *_init.xlsx
            input_files = sorted([
                f for f in files
                if f.endswith("_init.xlsx")
            ])

        if not input_files:
            # Try exact match for simple naming: initial.xlsx or input.xlsx
            for simple_name in ["initial.xlsx", "input.xlsx"]:
                if simple_name in files:
                    input_files = [simple_name]
                    break

        if not input_files:
            # Fall back to any xlsx that's not answer/output/golden
            input_files = sorted([
                f for f in files
                if f.endswith(".xlsx")
                and "answer" not in f.lower()
                and "output" not in f.lower()
                and "golden" not in f.lower()
            ])

        return input_files

    def run_instance(self, instance: BenchmarkInstance) -> InstanceResult:
        """Run the agent on a single benchmark instance."""
        spreadsheet_dir = self._find_spreadsheet_dir(instance)

        if spreadsheet_dir is None:
            return InstanceResult(
                id=instance.id,
                instruction=instance.instruction,
                success=False,
                error=f"Spreadsheet directory not found for {instance.id}",
            )

        input_files = self._find_input_files(spreadsheet_dir)

        if not input_files:
            return InstanceResult(
                id=instance.id,
                instruction=instance.instruction,
                success=False,
                error=f"No input files found in {spreadsheet_dir}",
            )

        result = InstanceResult(
            id=instance.id,
            instruction=instance.instruction,
            success=True,
        )

        for input_file in input_files:
            test_result = self._run_test_case(
                instance=instance,
                spreadsheet_dir=spreadsheet_dir,
                input_file=input_file,
            )
            result.test_cases.append(test_result)
            if not test_result.success:
                result.success = False

        return result

    def _run_test_case(
        self,
        instance: BenchmarkInstance,
        spreadsheet_dir: str,
        input_file: str,
    ) -> TestCaseResult:
        """Run a single test case."""
        input_path = os.path.join(spreadsheet_dir, input_file)

        # Determine output filename based on input naming convention
        if "_input.xlsx" in input_file:
            # Standard format: 1_Q001_input.xlsx -> 1_Q001_output.xlsx
            output_file = input_file.replace("_input.xlsx", "_output.xlsx")
        elif "_init.xlsx" in input_file:
            # Verified format: 1_13-1_init.xlsx -> 1_13-1_output.xlsx
            output_file = input_file.replace("_init.xlsx", "_output.xlsx")
        else:
            base = os.path.splitext(input_file)[0]
            output_file = f"{base}_output.xlsx"

        # Setup output path
        output_subdir = os.path.join(self.output_dir, instance.spreadsheet_path)
        os.makedirs(output_subdir, exist_ok=True)
        final_output_path = os.path.join(output_subdir, output_file)

        # Create task-specific subdirectory within working_dir
        # Use instance id and input file base name to create unique subdirectory
        input_base = os.path.splitext(input_file)[0]
        task_subdir_name = f"{instance.id}_{input_base}".replace("/", "_").replace("\\", "_")
        task_working_dir = os.path.join(self.working_dir, task_subdir_name)
        os.makedirs(task_working_dir, exist_ok=True)

        # Copy input to task-specific working directory
        work_input = os.path.join(task_working_dir, "input.xlsx")
        work_output = os.path.join(task_working_dir, "output.xlsx")
        shutil.copy(input_path, work_input)

        # Remove any existing output
        if os.path.exists(work_output):
            os.remove(work_output)

        # Get spreadsheet content
        spreadsheet_content = get_spreadsheet_content(work_input)

        # Create context
        context = AgentContext(
            working_dir=task_working_dir,
            input_file=work_input,
            output_file=work_output,
            instruction=instance.instruction,
            spreadsheet_content=spreadsheet_content,
            instruction_type=instance.instruction_type,
            answer_position=instance.answer_position,
            instance_id=instance.id,  # Use instance ID for log filename
        )

        # Run agent
        try:
            agent_result = self.agent.run(context)

            if os.path.exists(work_output):
                shutil.copy(work_output, final_output_path)
                return TestCaseResult(
                    input_file=input_file,
                    output_file=output_file,
                    success=True,
                    agent_answer=agent_result.get("answer", ""),
                    turns=agent_result.get("turns", 0),
                )
            else:
                return TestCaseResult(
                    input_file=input_file,
                    output_file=output_file,
                    success=False,
                    agent_answer=agent_result.get("answer", ""),
                    turns=agent_result.get("turns", 0),
                    error="Output file was not created",
                )

        except Exception as e:
            return TestCaseResult(
                input_file=input_file,
                output_file=output_file,
                success=False,
                error=f"{e}\n{traceback.format_exc()}",
            )

    def run(
        self,
        start_idx: int = 0,
        end_idx: int | None = None,
    ) -> BenchmarkResult:
        """
        Run the benchmark.

        Args:
            start_idx: Start index for instances
            end_idx: End index (exclusive) for instances

        Returns:
            BenchmarkResult with all results
        """
        # Load data
        instances = self.load_data()
        print(f"Loaded {len(instances)} instances")

        # Slice if needed
        end_idx = end_idx if end_idx is not None else len(instances)
        instances = instances[start_idx:end_idx]
        print(f"Running on instances {start_idx} to {end_idx} ({len(instances)} instances)")

        # Create output and working directories
        os.makedirs(self.output_dir, exist_ok=True)
        
        # Use custom working_dir if provided, otherwise create temp dir
        use_temp_dir = self._custom_working_dir is None
        if use_temp_dir:
            self.working_dir = tempfile.mkdtemp(prefix="spreadsheetbench_")
        else:
            self.working_dir = self._custom_working_dir
            os.makedirs(self.working_dir, exist_ok=True)
        print(f"Working directory: {self.working_dir}")

        try:
            results = []
            success_count = 0

            for i, instance in enumerate(instances):
                print(f"\n{'='*60}")
                print(f"Instance {i+1}/{len(instances)}: {instance.id}")
                print(f"Instruction: {instance.instruction[:100]}...")
                print(f"{'='*60}")

                result = self.run_instance(instance)
                results.append(result)

                if result.success:
                    success_count += 1

                print(f"Result: {'SUCCESS' if result.success else 'FAILED'}")
                if result.error:
                    print(f"Error: {result.error}")

            return BenchmarkResult(
                agent_name=self.agent.name,
                model=getattr(self.agent.client, "model", "unknown"),
                timestamp=datetime.now().isoformat(),
                total_instances=len(instances),
                successful_instances=success_count,
                success_rate=success_count / len(instances) if instances else 0,
                results=results,
            )

        finally:
            # Keep working_dir with task subdirectories for debugging/inspection
            # Each task has its own subdirectory that is preserved
            pass

    def run_and_save(
        self,
        results_file: str | None = None,
        start_idx: int = 0,
        end_idx: int | None = None,
    ) -> BenchmarkResult:
        """Run benchmark and save results to file."""
        result = self.run(start_idx=start_idx, end_idx=end_idx)

        # Save results
        results_file = results_file or os.path.join(self.output_dir, "results.json")

        # Convert to dict for JSON serialization
        result_dict = {
            "agent_name": result.agent_name,
            "model": result.model,
            "timestamp": result.timestamp,
            "total_instances": result.total_instances,
            "successful_instances": result.successful_instances,
            "success_rate": result.success_rate,
            "results": [
                {
                    "id": r.id,
                    "instruction": r.instruction,
                    "success": r.success,
                    "error": r.error,
                    "test_cases": [
                        {
                            "input_file": tc.input_file,
                            "output_file": tc.output_file,
                            "success": tc.success,
                            "agent_answer": tc.agent_answer,
                            "turns": tc.turns,
                            "error": tc.error,
                        }
                        for tc in r.test_cases
                    ],
                }
                for r in result.results
            ],
        }

        with open(results_file, "w") as f:
            json.dump(result_dict, f, indent=2)

        print(f"\n{'='*60}")
        print(f"BENCHMARK COMPLETE")
        print(f"Agent: {result.agent_name}")
        print(f"Total: {result.total_instances}, Success: {result.successful_instances}, "
              f"Rate: {result.success_rate*100:.1f}%")
        print(f"Results saved to: {results_file}")
        print(f"Outputs saved to: {self.output_dir}")
        print(f"{'='*60}")

        return result
