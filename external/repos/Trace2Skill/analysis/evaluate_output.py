#!/usr/bin/env python3
"""
Evaluation script for comparing an agent's output spreadsheet against ground truth.

Based on the official SpreadsheetBench evaluation logic (evaluation_official.py),
with rich diagnostic output so the error analysis agent can understand failures.

Usage:
    python evaluate_output.py --output_file path/to/output.xlsx \
                              --ground_truth path/to/gold.xlsx \
                              [--answer_position RANGE]
"""

import argparse
import datetime
import os
import sys

import openpyxl


# ---------------------------------------------------------------------------
# Value normalisation (from official SpreadsheetBench evaluation)
# ---------------------------------------------------------------------------

def datetime_to_float(dt):
    """Convert datetime to Excel serial-date float."""
    excel_start_date = datetime.datetime(1899, 12, 30)
    delta = dt - excel_start_date
    return delta.days + delta.seconds / 86400.0


def transform_value(v):
    """Normalise a cell value for comparison (official logic)."""
    if isinstance(v, (int, float)):
        v = round(float(v), 2)
    elif isinstance(v, datetime.time):
        v = str(v)[:-3]
    elif isinstance(v, datetime.datetime):
        v = round(datetime_to_float(v), 0)
    elif isinstance(v, str):
        try:
            v = round(float(v), 2)
        except ValueError:
            pass
    return v


def compare_cell_value(v1, v2):
    """Compare two cell values after normalisation (official logic)."""
    v1 = transform_value(v1)
    v2 = transform_value(v2)
    if (v1 == "" and v2 is None) or (v1 is None and v2 == ""):
        return True
    if (v1 == "" and v2 == "") or (v1 is None and v2 is None):
        return True
    if type(v1) != type(v2):
        return False
    return v1 == v2


# ---------------------------------------------------------------------------
# Cell range helpers (from official SpreadsheetBench evaluation)
# ---------------------------------------------------------------------------

def col_num2name(n):
    """Convert a 1-based column number to an Excel column name."""
    name = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        name = chr(65 + remainder) + name
    return name


def col_name2num(name):
    """Convert an Excel column name to a 1-based column number."""
    num = 0
    for c in name:
        num = num * 26 + (ord(c) - ord("A") + 1)
    return num


def parse_cell_range(range_str):
    """Parse a range string like 'A1:AB12' into ((col_num, row_num), (col_num, row_num))."""
    start_cell, end_cell = range_str.split(":")
    start_col, start_row = "", ""
    for char in start_cell:
        if char.isdigit():
            start_row += char
        else:
            start_col += char

    end_col, end_row = "", ""
    for char in end_cell:
        if char.isdigit():
            end_row += char
        else:
            end_col += char

    return (col_name2num(start_col), int(start_row)), (col_name2num(end_col), int(end_row))


def generate_cell_names(range_str):
    """Generate all cell names in a range (e.g. 'A1:B3' -> ['A1','A2','A3','B1','B2','B3'])."""
    if ":" not in range_str:
        return [range_str]
    (start_col, start_row), (end_col, end_row) = parse_cell_range(range_str)
    columns = [col_num2name(i) for i in range(start_col, end_col + 1)]
    return [f"{col}{row}" for col in columns for row in range(start_row, end_row + 1)]


# ---------------------------------------------------------------------------
# Cell-level comparison with rich diagnostics
# ---------------------------------------------------------------------------

def cell_level_compare(wb_gt, wb_proc, sheet_name, cell_range):
    """
    Compare cells in *cell_range* between two workbooks on *sheet_name*.

    Returns (passed: bool, message: str, mismatches: list[dict]).
    The mismatches list contains one entry per failing cell for diagnostics.
    """
    if sheet_name not in wb_proc.sheetnames:
        available = ", ".join(wb_proc.sheetnames) if wb_proc.sheetnames else "(none)"
        msg = (
            f"Sheet '{sheet_name}' not found in output workbook. "
            f"Available sheets: {available}"
        )
        return False, msg, []

    ws_gt = wb_gt[sheet_name]
    ws_proc = wb_proc[sheet_name]

    cell_names = generate_cell_names(cell_range)
    mismatches = []

    for cell_name in cell_names:
        cell_gt = ws_gt[cell_name]
        cell_proc = ws_proc[cell_name]

        if not compare_cell_value(cell_gt.value, cell_proc.value):
            gt_raw = cell_gt.value
            proc_raw = cell_proc.value
            gt_norm = transform_value(gt_raw)
            proc_norm = transform_value(proc_raw)
            mismatches.append({
                "cell": cell_name,
                "expected_raw": gt_raw,
                "actual_raw": proc_raw,
                "expected_normalized": gt_norm,
                "actual_normalized": proc_norm,
                "expected_type": type(gt_norm).__name__,
                "actual_type": type(proc_norm).__name__,
            })

    if mismatches:
        msg = f"{len(mismatches)} cell(s) differ in range {sheet_name}!{cell_range}"
        return False, msg, mismatches

    return True, "", []


# ---------------------------------------------------------------------------
# Top-level workbook comparison (matches official compare_workbooks signature)
# ---------------------------------------------------------------------------

def compare_workbooks(gt_file, proc_file, answer_position):
    """
    Compare *proc_file* against *gt_file* on the cells specified by *answer_position*.

    Returns (passed: bool, summary: str, details: list[dict]).
    """
    if not os.path.exists(proc_file):
        return False, f"Output file does not exist: {proc_file}", []
    if not os.path.exists(gt_file):
        return False, f"Ground truth file does not exist: {gt_file}", []

    try:
        wb_gt = openpyxl.load_workbook(filename=gt_file, data_only=True)
    except Exception as e:
        return False, f"Error loading ground truth workbook: {e}", []

    try:
        wb_proc = openpyxl.load_workbook(filename=proc_file, data_only=True)
    except Exception as e:
        return False, f"Error loading output workbook: {e}", []

    sheet_cell_ranges = answer_position.split(",")
    all_mismatches = []
    range_results = []

    for sheet_cell_range in sheet_cell_ranges:
        sheet_cell_range = sheet_cell_range.strip()
        if "!" in sheet_cell_range:
            sheet_name, cell_range = sheet_cell_range.split("!")
            sheet_name = sheet_name.strip("'")
        else:
            sheet_name = wb_gt.sheetnames[0]
            cell_range = sheet_cell_range

        sheet_name = sheet_name.strip("'")
        cell_range = cell_range.strip("'")

        passed, msg, mismatches = cell_level_compare(wb_gt, wb_proc, sheet_name, cell_range)
        range_results.append({
            "range": f"{sheet_name}!{cell_range}",
            "passed": passed,
            "message": msg,
            "mismatches": mismatches,
        })
        all_mismatches.extend(mismatches)

    overall_passed = all(r["passed"] for r in range_results)

    if overall_passed:
        summary = "PASS — all cells match."
    else:
        n_ranges_failed = sum(1 for r in range_results if not r["passed"])
        summary = (
            f"FAIL — {len(all_mismatches)} cell mismatch(es) across "
            f"{n_ranges_failed} range(s)."
        )

    return overall_passed, summary, range_results


# ---------------------------------------------------------------------------
# Pretty-print helpers
# ---------------------------------------------------------------------------

def _fmt_value(v):
    """Format a value for display, showing type for ambiguous cases."""
    if v is None:
        return "None (empty)"
    if isinstance(v, str):
        return f'"{v}" (str)'
    return f"{v} ({type(v).__name__})"


def print_report(passed, summary, range_results, output_file, gt_file, answer_position):
    """Print a human-readable evaluation report."""
    print("=" * 70)
    print("SPREADSHEET EVALUATION REPORT")
    print("=" * 70)
    print(f"Output file:     {output_file}")
    print(f"Ground truth:    {gt_file}")
    print(f"Answer position: {answer_position}")
    print(f"Result:          {'PASS' if passed else 'FAIL'}")
    print(f"Summary:         {summary}")
    print("-" * 70)

    if passed:
        print("All cells in the specified range(s) match the ground truth.")
        print("=" * 70)
        return

    for rr in range_results:
        range_label = rr["range"]
        if rr["passed"]:
            print(f"  [PASS] {range_label}")
            continue

        print(f"  [FAIL] {range_label}: {rr['message']}")
        mismatches = rr["mismatches"]
        if not mismatches:
            # Sheet-level error (e.g. missing sheet) — message already printed
            continue

        # Show up to 20 mismatched cells in detail
        show_limit = 20
        for i, m in enumerate(mismatches[:show_limit]):
            print(f"    Cell {m['cell']}:")
            print(f"      Expected : {_fmt_value(m['expected_raw'])}")
            print(f"      Got      : {_fmt_value(m['actual_raw'])}")
            if m["expected_type"] != m["actual_type"]:
                print(
                    f"      Type diff: expected {m['expected_type']}, "
                    f"got {m['actual_type']}"
                )
            elif m["expected_normalized"] != m["actual_normalized"]:
                print(
                    f"      Normalized: expected {m['expected_normalized']}, "
                    f"got {m['actual_normalized']}"
                )
        if len(mismatches) > show_limit:
            print(f"    ... and {len(mismatches) - show_limit} more mismatched cell(s)")

    print("-" * 70)
    total_mismatches = sum(len(rr["mismatches"]) for rr in range_results)
    print(f"Total mismatched cells: {total_mismatches}")
    print("=" * 70)


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Evaluate agent output against ground truth spreadsheet "
        "(SpreadsheetBench-compatible)."
    )
    parser.add_argument(
        "--output_file", required=True,
        help="Path to the agent's output .xlsx file",
    )
    parser.add_argument(
        "--ground_truth", required=True,
        help="Path to the ground-truth .xlsx file",
    )
    parser.add_argument(
        "--answer_position", default=None,
        help="Cell range(s) to compare, e.g. 'Sheet1!A1:B10' or 'A1:C5'. "
             "If omitted, compares all cells in every sheet.",
    )

    args = parser.parse_args()

    # If no answer_position provided, build one covering every sheet's used range
    answer_position = args.answer_position
    if not answer_position:
        if not os.path.exists(args.ground_truth):
            print(f"Error: ground truth file not found: {args.ground_truth}")
            sys.exit(1)
        try:
            wb = openpyxl.load_workbook(args.ground_truth, data_only=True)
        except Exception as e:
            print(f"Error loading ground truth: {e}")
            sys.exit(1)
        ranges = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row and ws.max_column:
                end_col = col_num2name(ws.max_column)
                ranges.append(f"'{sheet_name}'!A1:{end_col}{ws.max_row}")
        answer_position = ",".join(ranges) if ranges else "A1:Z100"
        print(f"No --answer_position given; comparing full sheets: {answer_position}")

    passed, summary, range_results = compare_workbooks(
        args.ground_truth, args.output_file, answer_position,
    )
    print_report(passed, summary, range_results, args.output_file, args.ground_truth, answer_position)

    sys.exit(0 if passed else 1)


if __name__ == "__main__":
    main()
